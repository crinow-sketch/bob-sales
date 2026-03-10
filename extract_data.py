"""Extract data from BOB brewery spreadsheets and create seed.json for the sales app."""
import json
import openpyxl
from datetime import datetime
import os
import uuid

BASE = r"C:\Users\crino\Desktop\Current Brewery Info"

def uid():
    return datetime.now().strftime("%y%m%d") + uuid.uuid4().hex[:8]

def safe_str(v):
    if v is None: return ""
    return str(v).strip()

def safe_int(v):
    if v is None: return 0
    try: return int(float(v))
    except: return 0

def safe_date(v):
    if v is None: return ""
    if isinstance(v, datetime):
        return v.strftime("%Y-%m-%d")
    s = str(v).strip()
    if not s: return ""
    # Try common formats
    for fmt in ["%Y-%m-%d", "%m/%d/%Y", "%m/%d/%y", "%m-%d-%Y"]:
        try:
            return datetime.strptime(s, fmt).strftime("%Y-%m-%d")
        except:
            pass
    return s

# ======== 1. Extract Account Database from CRM v2 ========
print("Reading CRM v2...")
wb = openpyxl.load_workbook(os.path.join(BASE, "BOB Sales & Marketing CRM (version 2).xlsx"), data_only=True)

accounts = []
seen_names = set()

# Account Database sheet
ws = wb["Account Database"]
rows = list(ws.iter_rows(min_row=2, values_only=True))
for row in rows:
    if not row or not row[0]:
        continue
    name = safe_str(row[0]).upper()
    if not name or name in seen_names:
        continue
    seen_names.add(name)

    acct = {
        "id": uid(),
        "name": safe_str(row[0]),
        "address": safe_str(row[1]) if len(row) > 1 else "",
        "city": safe_str(row[2]) if len(row) > 2 else "",
        "volumeRating": safe_int(row[3]) if len(row) > 3 else 0,
        "likelihoodRating": safe_int(row[4]) if len(row) > 4 else 0,
        "contactPerson": safe_str(row[5]) if len(row) > 5 else "",
        "lastVisitDate": safe_date(row[6]) if len(row) > 6 else "",
        "lastOrderDate": safe_date(row[7]) if len(row) > 7 else "",
        "productFormat": [safe_str(row[8])] if len(row) > 8 and row[8] else [],
        "status": safe_str(row[9]) if len(row) > 9 and row[9] else "Active",
        "notes": safe_str(row[10]) if len(row) > 10 else "",
        "contactPhone": "",
        "contactEmail": "",
        "onTap": "",
    }
    acct["priorityScore"] = acct["volumeRating"] + acct["likelihoodRating"]

    # Don't assume on-tap from order history - will be set from routing sheet below

    acct["createdAt"] = datetime.now().isoformat()
    acct["updatedAt"] = datetime.now().isoformat()
    accounts.append(acct)

print(f"  Account Database: {len(accounts)} accounts")

# ======== 2. Add accounts from Outside Accounts ========
print("Reading Outside Accounts...")
wb2 = openpyxl.load_workbook(os.path.join(BASE, "Outside Accounts.xlsx"), data_only=True)
ws2 = wb2["Accounts"]
rows2 = list(ws2.iter_rows(min_row=2, values_only=True))
added_outside = 0
for row in rows2:
    if not row or not row[0]:
        continue
    name = safe_str(row[0]).upper()
    if not name or name in seen_names:
        continue
    # Skip section headers
    if "HAVE PUT" in name or "HAVE VISITED" in name or "ACCOUNT" in name:
        continue
    seen_names.add(name)

    acct = {
        "id": uid(),
        "name": safe_str(row[0]),
        "address": safe_str(row[1]) if len(row) > 1 else "",
        "city": safe_str(row[2]) if len(row) > 2 else "",
        "volumeRating": safe_int(row[3]) if len(row) > 3 else 0,
        "likelihoodRating": safe_int(row[4]) if len(row) > 4 else 0,
        "contactPerson": "",
        "contactPhone": "",
        "contactEmail": "",
        "lastVisitDate": "",
        "lastOrderDate": "",
        "productFormat": [],
        "status": "Prospect",
        "notes": safe_str(row[5]) if len(row) > 5 else "",
        "onTap": "",
        "createdAt": datetime.now().isoformat(),
        "updatedAt": datetime.now().isoformat(),
    }
    acct["priorityScore"] = acct["volumeRating"] + acct["likelihoodRating"]
    accounts.append(acct)
    added_outside += 1

print(f"  Outside Accounts: +{added_outside} new accounts (total: {len(accounts)})")

# Build name->id lookup (case-insensitive, with fuzzy fallback)
name_to_id = {}
for a in accounts:
    name_to_id[a["name"].upper()] = a["id"]
    # Also add simplified versions (no punctuation)
    simplified = a["name"].upper().replace("'", "").replace("'", "").replace(".", "").replace(",", "").strip()
    name_to_id[simplified] = a["id"]

def find_account_id(name):
    """Find account ID with case-insensitive and fuzzy matching."""
    if not name:
        return ""
    upper = name.upper().strip()
    if upper in name_to_id:
        return name_to_id[upper]
    # Try without punctuation
    simplified = upper.replace("'", "").replace("'", "").replace(".", "").replace(",", "").strip()
    if simplified in name_to_id:
        return name_to_id[simplified]
    # Try substring match
    for key, aid in name_to_id.items():
        if upper in key or key in upper:
            return aid
    return ""

# ======== 3. Extract Activity Log ========
print("Reading Activity Log...")
activities = []
ws_act = wb["Activity Log"]
rows_act = list(ws_act.iter_rows(min_row=2, values_only=True))
for row in rows_act:
    if not row or not row[0]:
        continue
    acct_name = safe_str(row[1]).upper() if len(row) > 1 else ""
    acct_id = find_account_id(acct_name)

    act = {
        "id": uid(),
        "accountId": acct_id,
        "visitDate": safe_date(row[0]),
        "contactName": safe_str(row[3]) if len(row) > 3 else "",
        "interestLevel": safe_str(row[5]) if len(row) > 5 else "",
        "outcome": "",
        "visitNotes": safe_str(row[6]) if len(row) > 6 else "",
        "followUpAction": safe_str(row[7]) if len(row) > 7 else "",
        "followUpDate": safe_date(row[8]) if len(row) > 8 else "",
        "kegsOrdered": 0,
        "productsDiscussed": [],
        "createdAt": datetime.now().isoformat(),
        "updatedAt": datetime.now().isoformat(),
    }
    # Map interest to outcome
    interest = act["interestLevel"]
    if interest == "Hot":
        act["outcome"] = "Strong Interest"
    elif interest == "Warm":
        act["outcome"] = "Moderate Interest"
    elif interest == "Cold":
        act["outcome"] = "No Interest"
    elif interest in ("On Tap", "Ordered"):
        act["outcome"] = "Converted"

    activities.append(act)

print(f"  Activity Log: {len(activities)} entries")

# ======== 4. Also add Visit Log from Routing tracker ========
print("Reading Routing CRM Visit Log...")
wb3 = openpyxl.load_workbook(os.path.join(BASE, "BOB Sales Routing & CRM Tracker.xlsx"), data_only=True)
if "Visit Log" in wb3.sheetnames:
    ws_vl = wb3["Visit Log"]
    rows_vl = list(ws_vl.iter_rows(min_row=2, values_only=True))
    added_visits = 0
    for row in rows_vl:
        if not row or not row[0]:
            continue
        # Skip header/instruction rows (non-date values in first column)
        if not isinstance(row[0], datetime):
            test = safe_str(row[0])
            if not test or not test[0].isdigit():
                continue
        acct_name = safe_str(row[1]).upper() if len(row) > 1 else ""
        acct_id = find_account_id(acct_name)

        outcome = safe_str(row[4]) if len(row) > 4 else ""
        products = safe_str(row[5]) if len(row) > 5 else ""
        kegs = safe_int(row[6]) if len(row) > 6 else 0

        # Map outcome to interest level
        interest_map = {
            "Converted": "On Tap",
            "Strong Interest": "Hot",
            "Moderate Interest": "Warm",
            "No Interest": "Cold",
            "Left Sample": "Warm",
            "No Contact": "Cold",
        }

        act = {
            "id": uid(),
            "accountId": acct_id,
            "visitDate": safe_date(row[0]),
            "contactName": safe_str(row[3]) if len(row) > 3 else "",
            "interestLevel": interest_map.get(outcome, ""),
            "outcome": outcome,
            "visitNotes": safe_str(row[9]) if len(row) > 9 else "",
            "followUpAction": safe_str(row[8]) if len(row) > 8 else "",
            "followUpDate": safe_date(row[9]) if len(row) > 9 and isinstance(row[9], datetime) else "",
            "kegsOrdered": kegs,
            "productsDiscussed": [products] if products else [],
            "createdAt": datetime.now().isoformat(),
            "updatedAt": datetime.now().isoformat(),
        }
        activities.append(act)
        added_visits += 1
    print(f"  Visit Log: +{added_visits} entries (total activities: {len(activities)})")

# ======== 5. Extract Pipeline ========
print("Reading Pipeline...")
pipeline = []
ws_pipe = wb["Pipeline"]
rows_pipe = list(ws_pipe.iter_rows(min_row=2, values_only=True))
for row in rows_pipe:
    if not row or not row[0]:
        continue
    acct_name = safe_str(row[0]).upper()
    acct_id = find_account_id(acct_name)

    stage = safe_str(row[4]) if len(row) > 4 else "Cold"
    # Normalize stage names
    stage_map = {
        "Warm Prospect": "Warm Prospect",
        "Hot Lead": "Hot Lead",
        "Needs Follow-Up": "Needs Follow-Up",
        "On Tap": "On Tap",
        "Ordered": "Ordered",
        "Cold": "Cold",
        "Cold Lead": "Cold",
    }
    stage = stage_map.get(stage, stage)

    p = {
        "id": uid(),
        "accountId": acct_id,
        "stage": stage,
        "priority": safe_str(row[5]) if len(row) > 5 else "Medium",
        "targetProducts": safe_str(row[8]) if len(row) > 8 else "Both",
        "lastContactDate": safe_date(row[6]) if len(row) > 6 else "",
        "nextAction": safe_str(row[7]) if len(row) > 7 else "",
        "notes": safe_str(row[9]) if len(row) > 9 else "",
        "createdAt": datetime.now().isoformat(),
        "updatedAt": datetime.now().isoformat(),
    }
    pipeline.append(p)

print(f"  Pipeline: {len(pipeline)} entries")

# ======== 6. Extract Sales Performance ========
print("Reading Sales Performance...")
sales = []
ws_sales = wb["Sales Performance"]
rows_sales = list(ws_sales.iter_rows(min_row=2, values_only=True))
for row in rows_sales:
    if not row or not row[0]:
        continue
    month = safe_str(row[0])
    if month in ("GOAL", "AVERAGE", "Total", ""):
        continue
    year = safe_int(row[1]) if len(row) > 1 else 0
    if not year:
        continue
    kegs = safe_int(row[2]) if len(row) > 2 else 0
    cases = safe_int(row[3]) if len(row) > 3 else 0
    notes = safe_str(row[4]) if len(row) > 4 else ""
    goal_pct = row[5] if len(row) > 5 and row[5] else 0

    # Estimate product breakdown (roughly 50/50 split IPA/Lager)
    ipa_kegs = kegs // 2
    lager_kegs = kegs - ipa_kegs

    s = {
        "id": uid(),
        "month": month,
        "year": year,
        "kegsIPA_sixth": ipa_kegs // 2 if ipa_kegs > 1 else ipa_kegs,
        "kegsIPA_half": ipa_kegs - (ipa_kegs // 2 if ipa_kegs > 1 else ipa_kegs),
        "kegsLager_sixth": lager_kegs // 2 if lager_kegs > 1 else lager_kegs,
        "kegsLager_half": lager_kegs - (lager_kegs // 2 if lager_kegs > 1 else lager_kegs),
        "kegsSold": kegs,
        "casesCans": cases,
        "goal": 20,  # Default goal
        "notes": notes,
        "createdAt": datetime.now().isoformat(),
        "updatedAt": datetime.now().isoformat(),
    }
    sales.append(s)

print(f"  Sales Performance: {len(sales)} months")

# ======== 7. Extract Routes from 3-Week Routing ========
print("Reading 3-Week Routing...")
routes = []
if "3-Week Routing" in wb3.sheetnames:
    ws_route = wb3["3-Week Routing"]
    rows_route = list(ws_route.iter_rows(min_row=2, values_only=True))
    for row in rows_route:
        if not row or not row[3]:  # Account Name is col D (index 3)
            continue
        acct_name = safe_str(row[3]).upper()
        acct_id = find_account_id(acct_name)

        week = safe_int(row[0]) if row[0] else 0
        day = safe_str(row[1]) if len(row) > 1 else ""
        stop = safe_int(row[2]) if len(row) > 2 else 0

        if not week or not day:
            continue

        outcome = safe_str(row[13]) if len(row) > 13 else ""
        on_tap = safe_str(row[14]) if len(row) > 14 else ""
        kegs = safe_int(row[15]) if len(row) > 15 else 0
        contact_made = safe_str(row[12]) if len(row) > 12 else ""

        r = {
            "id": uid(),
            "accountId": acct_id,
            "week": week,
            "day": day,
            "stopNumber": stop,
            "hours": safe_str(row[6]) if len(row) > 6 else "",
            "contactMade": contact_made,
            "outcome": outcome,
            "kegsOrdered": kegs,
            "followUpDate": safe_date(row[17]) if len(row) > 17 else "",
            "notes": safe_str(row[16]) if len(row) > 16 else "",
            "createdAt": datetime.now().isoformat(),
            "updatedAt": datetime.now().isoformat(),
        }
        routes.append(r)

        # Update the account's on-tap status from routing data
        if on_tap and acct_id:
            on_tap_val = on_tap.strip()
            if on_tap_val in ("Yes", "No", "Pending"):
                for acct in accounts:
                    if acct["id"] == acct_id:
                        acct["onTap"] = on_tap_val
                        break

    # Also check visit log for "Converted" outcomes to mark on-tap
    for act in activities:
        if act["outcome"] == "Converted" and act["accountId"]:
            for acct in accounts:
                if acct["id"] == act["accountId"] and not acct["onTap"]:
                    acct["onTap"] = "Yes"
                    break

on_tap_count = sum(1 for a in accounts if a["onTap"] == "Yes")
print(f"  Routes: {len(routes)} stops")
print(f"  On Tap accounts: {on_tap_count}")

# ======== Build final seed data ========
seed = {
    "accounts": accounts,
    "activities": activities,
    "pipeline": pipeline,
    "sales": sales,
    "routes": routes,
    "exportedAt": datetime.now().isoformat(),
}

out_path = os.path.join(BASE, "sales-app", "data", "seed.json")
os.makedirs(os.path.dirname(out_path), exist_ok=True)
with open(out_path, "w", encoding="utf-8") as f:
    json.dump(seed, f, indent=2, ensure_ascii=False)

print(f"\nSeed file written to: {out_path}")
print(f"  Accounts: {len(accounts)}")
print(f"  Activities: {len(activities)}")
print(f"  Pipeline: {len(pipeline)}")
print(f"  Sales: {len(sales)}")
print(f"  Routes: {len(routes)}")
